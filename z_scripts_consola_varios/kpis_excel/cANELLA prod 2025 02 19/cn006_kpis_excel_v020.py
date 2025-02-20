import argparse
from collections import defaultdict
from datetime import datetime, timedelta
import json
import locale
import openpyxl
import os
import pytz
import sys
import xmlrpc.client



sys.path.append(os.path.abspath(".."))  # Agrega el directorio superior al path
from cn006_kpis_globales import cCN006_globales


#region EXPLICACIÓN DEL PROCESO
"""  ******* EXPLICACIÓN DEL PROCESO
    1. Obtener todos los proyectos de TI (cn006_project = True)
    2. Obtener todos los tickets de esos proyectos
    3. Combinar los tickets con los proyectos
************************************************************************  """
#endregion

######################################################################################################
#region GLOBALES
    ######################################################################################################
    #region Variables tipo fecha
COLUMNAS_FECHA = {
    "ph_date", "proy_cn006_fecha_cierre_estimada", "proy_cn006_fecha_cierre_oficial",
    "proy_cn006_fecha_cierre_sistema", "proy_cn006_fecha_creacion_oficial",
    "proy_cn006_fecha_creacion_sistema", "proy_cn006_fecha_entrega_informatica_estimada",
    "proy_cn006_fecha_entrega_informatica_oficial", "proy_cn006_fecha_entrega_informatica_sistema",
    "proy_cn006_fecha_entrega_usuario_estimada", "proy_cn006_fecha_entrega_usuario_oficial",
    "proy_cn006_fecha_entrega_usuario_sistema", "proy_cn006_fecha_inicio_oficial",
    'proy_cn006_fecha_gerencia_estimada','proy_cn006_fecha_gerencia_oficial','proy_cn006_fecha_gerencia_sistema',
    "proy_cn006_fecha_inicio_sistema", "proy_date", "proy_date_start",
    'Fecha Creación', 'Fecha Inicio', 'Fecha Informática', 'Fecha Gerencia', 'Fecha Cierre'
}

COLUMNAS_FECHA_HORA = {
    "ph_create_date", "ph_write_date", "proy_create_date", "proy_write_date"
}

DIAS_SEMANA = {1: "LUN", 2: "MAR", 3: "MIE", 4: "JUE", 5: "VIE", 6: "SAB", 7: "DOM"}
MESES = {
    1: "ENE", 2: "FEB", 3: "MAR", 4: "ABR", 5: "MAY", 6: "JUN",
    7: "JUL", 8: "AGO", 9: "SEP", 10: "OCT", 11: "NOV", 12: "DIC"
}
    #endregion Variables tipo fecha
    ######################################################################################################
#endregion GLOBALES
######################################################################################################

def msj_debug (p_msj, p_tools: cCN006_globales):
    if p_tools.g_debug:
        print(p_msj)
    return

################################################################################################################
# Extracción de información
################################################################################################################

######################################################################################################
#region OBTENER PROYECTOS
def obtener_proyectos(p_tools: cCN006_globales):
    models = xmlrpc.client.ServerProxy(f"{p_tools.cnx_url}/xmlrpc/2/object")

    # Obtener proyectos donde cn006_project = True
    # agregar el campo cn006_emergente
    proyectos = models.execute_kw(
                                    p_tools.cnx_db, p_tools.cnx_uid, p_tools.cnx_password,
                                    'project.project', 'search_read', 
                                    [[('cn006_project', '=', True)]], 
                                    {'fields': [
                                        'id', 'name', 'user_id', 'active', 
                                        'company_id', 'partner_id', 'date_start', 'date',
                                        'description',
                                        'cn006_clasificacion_id', 
                                        'stage_id', 'cn006_grado_complejidad_id', 'cn006_nivel_importancia_id', 
                                        'cn006_nivel_urgencia_id', 'cn006_tamano_id', 'cn006_emergente','cn006_stod_codigo',
                                        'cn006_project',
                                        'cn006_fecha_creacion_sistema', 'cn006_fecha_creacion_oficial', 'cn006_fecha_inicio_oficial', 'cn006_fecha_inicio_sistema', 
                                        'cn006_fecha_entrega_informatica_estimada', 'cn006_fecha_entrega_informatica_oficial', 'cn006_fecha_entrega_informatica_sistema', 
                                        'cn006_fecha_entrega_usuario_estimada', 'cn006_fecha_entrega_usuario_oficial', 'cn006_fecha_entrega_usuario_sistema', 
                                        'cn006_fecha_cierre_estimada', 'cn006_fecha_cierre_oficial', 'cn006_fecha_cierre_sistema',
                                        'cn006_fecha_gerencia_estimada','cn006_fecha_gerencia_oficial','cn006_fecha_gerencia_sistema',
                                        'create_date', 'create_uid', 'write_date', 'write_uid',
                                        'timesheet_ids', 'task_ids'
                                        ],
                                        'context': {'lang': 'es_GT'} 
                                    }
                                )
    
    # Ajustando los campos para que sean manejables en excel
    for proyecto in proyectos:
        #region Campos que son estructura
        if proyecto['user_id']:
            proyecto['user_id_name'] = proyecto['user_id'][1]
            proyecto['user_id'] = proyecto['user_id'][0]
        else:
            proyecto['user_id_name'] = None
            proyecto['user_id'] = None

        if proyecto['company_id']:
            proyecto['company_id_name'] = proyecto['company_id'][1]
            proyecto['company_id'] = proyecto['company_id'][0]
        else:
            proyecto['company_id_name'] = None
            proyecto['company_id'] = None

        if proyecto['partner_id']:
            proyecto['partner_id_name'] = proyecto['partner_id'][1]
            proyecto['partner_id'] = proyecto['partner_id'][0]
        else:
            proyecto['partner_id_name'] = None
            proyecto['partner_id'] = None

        if proyecto['stage_id']:
            proyecto['stage_id_name'] = proyecto['stage_id'][1]
            proyecto['stage_id'] = proyecto['stage_id'][0]
        else:
            proyecto['stage_id_name'] = None
            proyecto['stage_id'] = None
        
        if proyecto['cn006_clasificacion_id']:
            proyecto['cn006_clasificacion_id_name'] = proyecto['cn006_clasificacion_id'][1]
            proyecto['cn006_clasificacion_id'] = proyecto['cn006_clasificacion_id'][0]
        else:
            proyecto['cn006_clasificacion_id_name'] = None
            proyecto['cn006_clasificacion_id'] = None

        if proyecto['cn006_grado_complejidad_id']:
            proyecto['cn006_grado_complejidad_name'] = proyecto['cn006_grado_complejidad_id'][1]
            proyecto['cn006_grado_complejidad_id'] = proyecto['cn006_grado_complejidad_id'][0]
        else:
            proyecto['cn006_grado_complejidad_id_name'] = None
            proyecto['cn006_grado_complejidad_id'] = None
        
        if proyecto['cn006_nivel_importancia_id']:
            proyecto['cn006_nivel_importancia_id_name'] = proyecto['cn006_nivel_importancia_id'][1]
            proyecto['cn006_nivel_importancia_id'] = proyecto['cn006_nivel_importancia_id'][0]
        else:
            proyecto['cn006_nivel_importancia_id_name'] = None
            proyecto['cn006_nivel_importancia_id'] = None

        if proyecto['cn006_nivel_urgencia_id']:
            proyecto['cn006_nivel_urgencia_id_name'] = proyecto['cn006_nivel_urgencia_id'][1]
            proyecto['cn006_nivel_urgencia_id'] = proyecto['cn006_nivel_urgencia_id'][0]
        else:
            proyecto['cn006_nivel_urgencia_id_name'] = None
            proyecto['cn006_nivel_urgencia_id'] = None

        if proyecto['cn006_tamano_id']:
            proyecto['cn006_tamano_id_name'] = proyecto['cn006_tamano_id'][1]
            proyecto['cn006_tamano_id'] = proyecto['cn006_tamano_id'][0]
        else:
            proyecto['cn006_tamano_id_name'] = None
            proyecto['cn006_tamano_id'] = None

        if proyecto['create_uid']:
            proyecto['create_uid_name'] = proyecto['create_uid'][1]
            proyecto['create_uid'] = proyecto['create_uid'][0]
        else:
            proyecto['create_uid_name'] = None
            proyecto['create_uid'] = None

        if proyecto['write_uid']:
            proyecto['write_uid_name'] = proyecto['write_uid'][1]
            proyecto['write_uid'] = proyecto['write_uid'][0]
        else:
            proyecto['write_uid_name'] = None
            proyecto['write_uid'] = None


        if proyecto['task_ids']:
            proyecto['task_ids'] = ", ".join(map(str, proyecto['task_ids']  ))
        else:
            proyecto['task_ids'] = None

        #endregion Campos que son estructura
    # Agregando prefijo proy_
    proyectos = [
        {f'proy_{key}': value for key, value in detalle.items()}
        for detalle in proyectos
        ]  

    return proyectos
#endregion OBTENER PROYECTOS
######################################################################################################

######################################################################################################
#region OBTENER PARTES DE HORAS
def obtener_partes_horas(p_proyectos, p_tools: cCN006_globales):
    partes_horas_proyectos = []
    for proyecto in p_proyectos:
        partes_horas_proyectos.extend(proyecto.get('proy_timesheet_ids', []))

    msj_debug(f"En total hay ({len(partes_horas_proyectos)}) partes de horas de proyectos\n", p_tools)
    

    if not partes_horas_proyectos:
        print("No hay partes de horas para consultar.")
        return []


    msj_debug("Previo a consulta de detalle partes de horas en la base de datos", p_tools)
    detalles_partes_horas = []
    models = xmlrpc.client.ServerProxy(f"{p_tools.cnx_url}/xmlrpc/2/object")
    detalles_partes_horas = models.execute_kw(
        p_tools.cnx_db, p_tools.cnx_uid, p_tools.cnx_password,
        'account.analytic.line', 'search_read',
        [[('id', 'in', partes_horas_proyectos)]],
        {'fields': ['project_id','id', 'name', 'date', 'unit_amount', 'amount', 
                    'user_id', 'account_id', 'task_id',
                    'create_date', 'create_uid', 'write_date', 'write_uid'
                    ]}

    )

    msj_debug("regreso de consulta de detalle de partes de horas en la base de datos", p_tools)

    # Ajustando los campos para que sean manejables en excel
    for detalle in detalles_partes_horas:
            if detalle['project_id']:
                detalle['project_id_name'] = detalle['project_id'][1]
                detalle['project_id'] = detalle['project_id'][0]
            else:
                detalle['project_id_name'] = None
                detalle['project_id'] = None

            if detalle['user_id']:
                detalle['user_id_name'] = detalle['user_id'][1]
                detalle['user_id'] = detalle['user_id'][0]
            else:
                detalle['user_id_name'] = None
                detalle['user_id'] = None

            if detalle['account_id']:
                detalle['account_id_name'] = detalle['account_id'][1]
                detalle['account_id'] = detalle['account_id'][0]
            else:
                detalle['account_id_name'] = None
                detalle['account_id'] = None

            if detalle['task_id']:
                detalle['task_id_name'] = detalle['task_id'][1]
                detalle['task_id'] = detalle['task_id'][0]
            else:
                detalle['task_id_name'] = None
                detalle['task_id'] = None

            if detalle['create_uid']:
                detalle['create_uid_name'] = detalle['create_uid'][1]
                detalle['create_uid'] = detalle['create_uid'][0]
            else:
                detalle['create_uid_name'] = None
                detalle['create_uid_id'] = None

            if detalle['write_uid']:
                detalle['write_uid_name'] = detalle['write_uid'][1]
                detalle['write_uid'] = detalle['write_uid'][0]
            else:
                detalle['write_uid_name'] = None
                detalle['write_uid_id'] = None
    
    # Agregando prefijo ph_
    detalles_partes_horas = [
        {f'ph_{key}': value for key, value in detalle.items()}
        for detalle in detalles_partes_horas
        ]   
    # msj_debug("Estos son los registros de partes de horas", p_tools)
    # for detalle in detalles_partes_horas:
    #     print(f"\n{detalle}")
    return detalles_partes_horas
#endregion OBTENER PARTES DE HORAS
######################################################################################################

######################################################################################################
#region OBTENER TAREAS
def obtener_tareas(p_partes_horas, p_tools: cCN006_globales):
    msj_debug("\n\n*****  INICIO obtener_tareas", p_tools)
    msj_debug(f"En total hay ({len(p_partes_horas)}) partes de horas de proyectos para obtener tareas\n", p_tools)
    
    if not p_partes_horas:
        msj_debug("No hay partes de horas para consultar.", p_tools)
        return []
    
    # print("Para verificación, estos son (5) registros de partes de horas DENTRO DE OBTENER TAREAS")
    # contador = 0
    # for parte_horas in p_partes_horas:
    #     contador += 1
    #     if contador <= 5:
    #         print(f"Parte de horas: {parte_horas}\n")    

    task_ids = list(set(p['ph_task_id'] for p in p_partes_horas if p.get('ph_task_id')))

    if not task_ids:
        msj_debug("No hay tareas asociadas a los partes de horas.", p_tools)
        return p_partes_horas
    
    msj_debug(f"Consultando {len(task_ids)} tareas en la base de datos.", p_tools)
    
    msj_debug("Previo a consulta de TAREAS en la base de datos", p_tools)
    models = xmlrpc.client.ServerProxy(f"{p_tools.cnx_url}/xmlrpc/2/object")
    tareas = models.execute_kw(
        p_tools.cnx_db, p_tools.cnx_uid, p_tools.cnx_password,
        'project.task', 'read', [task_ids],
        {'fields': ['analytic_account_id', 'cn006_es_implementacion', 'cn006_grado_avance_id', 
                    'cn006_tipificacion_id', 'cn006_tipo_soporte_id', 'stage_id']}
    )
    
    msj_debug(f"La base de datos retornó ({len(tareas)}) tareas asociadas a los partes de horas.", p_tools)

    tarea_dict = {t['id']: t for t in tareas}

    # print(f"Esto es tarea_dict:\n")
    # print(json.dumps(tarea_dict, indent=4, ensure_ascii=False))
    # print("\n\n")

    msj_debug("Previo al ciclo de análisis de partes de horas", p_tools)
    for parte in p_partes_horas:
        #task_id = parte.get('task_id')

        #print(f"Parte de horas antes de asignar tarea: {parte}")
        task_id = parte.get('ph_task_id') if parte.get('ph_task_id') else None
        #print(f"Task ID obtenido: {task_id}")

        if task_id and task_id in tarea_dict:
            tarea = tarea_dict[task_id]
            #print(f"Tarea encontrada: {tarea}")
            parte['tarea_analytic_account_id'] = tarea.get('analytic_account_id', [False])[0] if isinstance(tarea.get('analytic_account_id'), list) else tarea.get('analytic_account_id')
            parte['tarea_analytic_account_id_name'] = tarea.get('analytic_account_id', [False])[1] if isinstance(tarea.get('analytic_account_id'), list) else ""
            
            parte['tarea_cn006_es_implementacion'] = tarea.get('cn006_es_implementacion', False)
            parte['tarea_cn006_es_implementacion_name'] = "Sí" if tarea.get('cn006_es_implementacion') else "No"
            
            parte['tarea_cn006_grado_avance_id'] = tarea.get('cn006_grado_avance_id', [False])[0] if isinstance(tarea.get('cn006_grado_avance_id'), list) else tarea.get('cn006_grado_avance_id')
            parte['tarea_cn006_grado_avance_id_name'] = tarea.get('cn006_grado_avance_id', [False])[1] if isinstance(tarea.get('cn006_grado_avance_id'), list) else ""
            
            parte['tarea_cn006_tipificacion_id'] = tarea.get('cn006_tipificacion_id', [False])[0] if isinstance(tarea.get('cn006_tipificacion_id'), list) else tarea.get('cn006_tipificacion_id')
            parte['tarea_cn006_tipificacion_id_name'] = tarea.get('cn006_tipificacion_id', [False])[1] if isinstance(tarea.get('cn006_tipificacion_id'), list) else ""
            
            parte['tarea_cn006_tipo_soporte_id'] = tarea.get('cn006_tipo_soporte_id', [False])[0] if isinstance(tarea.get('cn006_tipo_soporte_id'), list) else tarea.get('cn006_tipo_soporte_id')
            parte['tarea_cn006_tipo_soporte_id_name'] = tarea.get('cn006_tipo_soporte_id', [False])[1] if isinstance(tarea.get('cn006_tipo_soporte_id'), list) else ""
            
            parte['tarea_stage_id'] = tarea.get('stage_id', [False])[0] if isinstance(tarea.get('stage_id'), list) else tarea.get('stage_id')
            parte['tarea_stage_id_name'] = tarea.get('stage_id', [False])[1] if isinstance(tarea.get('stage_id'), list) else ""
            
            #msj_debug(f"Tarea {task_id} procesada y asignada a parte de horas.", p_tools)
    
    msj_debug("Proceso de actualización de partes de horas completado.", p_tools)

    # print(f"Estos son los (5) registros de partes de horas con las tareas asociadas DENTRO DE OBTENER TAREAS")
    # contador = 0
    # for detalle in p_partes_horas:
    #     contador += 1
    #     if contador <= 5:
    #         print(f"\n{detalle}")   

    msj_debug("\n\n*****  FINALIZO obtener_tareas", p_tools)
    return p_partes_horas
    
#endregion OBTENER TAREAS
######################################################################################################

######################################################################################################
#region UNIFICAR DATOS PROYECTOS + PARTES DE HORAS +  TASKS
def unificar_datos(p_proyectos, p_partes_horas,p_tools: cCN006_globales):
    proyectos_info_total = []
    for proyecto in p_proyectos:
        #partes_horas = [parte for parte in p_partes_horas if parte["ph_project_id"][0] == proyecto["proy_id"]]
        partes_horas = [parte for parte in p_partes_horas if parte["ph_project_id"] == proyecto["proy_id"]]

        # Escenario 1: Proyecto sin partes de horas
        if not partes_horas:
            proyectos_info_total.append({**proyecto, 
                                            "ph_project_id": None, 
                                            "ph_id": None, 
                                            "ph_name": None, 
                                            "ph_date": None,
                                            "ph_unit_amount": None, 
                                            "ph_amount": None,
                                            "ph_user_id": None, 
                                            "ph_account_id": None, 
                                            "ph_task_id": None
                                        }
                                            )

        # Escenario 2 y 3: Proyecto con una o más partes de horas
        else:
            for parte in partes_horas:
                parte_convertida = parte.copy()
                # Conversión de horas decimales a formato de tiempo para Excel
                if parte_convertida.get("ph_unit_amount") is not None:
                    parte_convertida["ph_unit_amount"] = parte_convertida["ph_unit_amount"] / 24
                
                # msj_debug(f"Parte con ID {parte_convertida['ph_id']} tiene las siguientes claves: {parte_convertida.keys()}", tools)


                proyectos_info_total.append({**proyecto, **parte_convertida})
    return proyectos_info_total

#endregion UNIFICAR DATOS PROYECTOS + PARTES DE HORAS +  TASKS
######################################################################################################

######################################################################################################
#region AJUSTES FINALES DATA
def _desglose_fecha(proyecto, campo, fecha):
    """Agrega al diccionario los campos desglosados de la fecha."""
    # Establecer el locale a español para los nombres de los meses
    locale.setlocale(locale.LC_TIME, 'es_ES.utf8')
    if fecha is None:
        proyecto[f"{campo}_sem_txt"] = ""
        proyecto[f"{campo}_sem"] = 0
        proyecto[f"{campo}_dd"] = 0
        proyecto[f"{campo}_dd_nombre"] = ""
        proyecto[f"{campo}_mm"] = 0
        proyecto[f"{campo}_mm_nombre"] = ""
        proyecto[f"{campo}_yy"] = 0
        proyecto[f"{campo}_yyyy"] = 0
    else:
         # Calcular número de semana ISO
        numero_semana = fecha.isocalendar().week

        # Calcular el lunes de esa semana
        lunes_semana = fecha - timedelta(days=fecha.weekday())
        lunes_formateado = lunes_semana.strftime("%d-%b").lower()

        proyecto[f"{campo}_sem_txt"] = f"({numero_semana:02d}) {lunes_formateado}"
        proyecto[f"{campo}_sem"] = fecha.isocalendar()[1]
        proyecto[f"{campo}_dd"] = fecha.isoweekday()
        proyecto[f"{campo}_dd_nombre"] = DIAS_SEMANA.get(fecha.isoweekday(), "")
        proyecto[f"{campo}_mm"] = fecha.month
        proyecto[f"{campo}_mm_nombre"] = MESES.get(fecha.month, "")
        proyecto[f"{campo}_yy"] = fecha.year % 100
        proyecto[f"{campo}_yyyy"] = fecha.year

def ajustes_finales_data(p_proyectos_info_total,p_tools: cCN006_globales):
    # Zona horaria UTC-6 (Guatemala)
    tz_gt = pytz.timezone("America/Guatemala")


    msj_debug("\n\n*****  INICIO ajustes_finales_data", p_tools)
    msj_debug(f"En total hay ({len(p_proyectos_info_total)}) registros de proyectos con partes de horas\n", p_tools)
    
    msj_debug("INICIANDO ciclo de revisión de datos", p_tools)

    proyectos_vistos = set()

    stage_order = {
            "COLA": "10-COLA",
            "ASIGNADO": "20-ASIGNADO",
            "ANALISIS": "30-ANALISIS",
            "DESARROLLO": "40-DESARROLLO",
            "IMPLEMENTACION": "50-IMPLEMENTACION",
            "CERTIFICADO": "60-CERTIFICADO"
        }
    for proyecto in p_proyectos_info_total:
        #############################################################
        #region Prefijos pendientes proyectos
        if proyecto['proy_timesheet_ids']:
            proyecto['proy_timesheet_ids'] = ", ".join(map(str, proyecto['proy_timesheet_ids']))
        else:
            proyecto['proy_timesheet_ids'] = None
        #endregion Prefijos pendientes proyectos
        #############################################################

        #############################################################
        #region procesar campos de fecha

        # Generar lista de campos a procesar para evitar modificar el diccionario mientras se itera
        campos_fecha = [campo for campo in proyecto.keys() if campo in COLUMNAS_FECHA or campo in COLUMNAS_FECHA_HORA]

        # Procesar campos de fecha
        for campo in campos_fecha:
            valor = proyecto[campo]

            # Si el valor es False o None, establecerlo como None y desglosar en ceros
            if not valor:
                proyecto[campo] = None
                _desglose_fecha(proyecto, campo, None)
                continue

            # Convertir string a datetime si es necesario
            if isinstance(valor, str):
                formato = "%Y-%m-%d %H:%M:%S" if campo in COLUMNAS_FECHA_HORA else "%Y-%m-%d"
                valor = datetime.strptime(valor, formato)

            # Si es fecha_hora, ajustar a UTC-6
            if campo in COLUMNAS_FECHA_HORA:
                valor = pytz.utc.localize(valor).astimezone(tz_gt)

            # Asignar el valor final al proyecto
            proyecto[campo] = valor.replace(tzinfo=None)

            # Desglosar la fecha en sus componentes
            _desglose_fecha(proyecto, campo, valor)

        #endregion procesar campos de fecha
        #############################################################

        #############################################################
        #region identificar proyectos únicos
        proy_id = proyecto.get('proy_id')
        if proy_id not in proyectos_vistos:
            proyecto['proy_id_unico'] = "x"
            proyectos_vistos.add(proy_id)
        else:
            proyecto['proy_id_unico'] = ""
        #endregion identificar proyectos únicos
        #############################################################

        stage = proyecto.get('proy_stage_id_name', '')
        proyecto['proy_stage_id_name'] = stage_order.get(stage, stage)


    msj_debug("FINALIZADO ciclo de revisión de datos", p_tools)    


    msj_debug("Ajustes finales de datos completados.", p_tools)
    msj_debug("\n\n*****  FINALIZO ajustes_finales_data", p_tools)
    return p_proyectos_info_total
#endregion AJUSTES FINALES DATA
######################################################################################################

######################################################################################################
#region SELECCIÓN DE DATOS PARA EXCEL
def seleccionar_campos_finales(p_proyectos_info_total, p_tools: cCN006_globales):
    # Estructura para definir los campos a seleccionar y sus nuevos nombres
    campos_seleccionados = [
        {'nombre_actual': 'proy_id', 'nombre_nuevo': 'Proyecto ID'},
        {'nombre_actual': 'proy_cn006_stod_codigo', 'nombre_nuevo': 'STOD Ref.'},
        {'nombre_actual': 'proy_name', 'nombre_nuevo': 'Proyecto'},
        {'nombre_actual': 'proy_partner_id', 'nombre_nuevo': 'Solicitante ID'},
        {'nombre_actual': 'proy_partner_id_name', 'nombre_nuevo': 'Solicitante'},
        {'nombre_actual': 'proy_stage_id', 'nombre_nuevo': 'Etapa ID'},
        {'nombre_actual': 'proy_stage_id_name', 'nombre_nuevo': 'Etapa'},
        {'nombre_actual': 'proy_user_id', 'nombre_nuevo': 'Asignado ID'},
        {'nombre_actual': 'proy_user_id_name', 'nombre_nuevo': 'Asignado'},
        {'nombre_actual': 'proy_id_unico', 'nombre_nuevo': 'Es único?'},
        {'nombre_actual': 'proy_cn006_emergente', 'nombre_nuevo': 'Emergente'},
        {'nombre_actual': 'proy_cn006_clasificacion_id', 'nombre_nuevo': 'Clasificación ID'},
        {'nombre_actual': 'proy_cn006_clasificacion_id_name', 'nombre_nuevo': 'Clasificación'},
        {'nombre_actual': 'proy_cn006_grado_complejidad_id', 'nombre_nuevo': 'Complejidad ID'},
        {'nombre_actual': 'proy_cn006_grado_complejidad_name', 'nombre_nuevo': 'Complejidad'},
        {'nombre_actual': 'proy_cn006_nivel_importancia_id', 'nombre_nuevo': 'Importancia ID'},
        {'nombre_actual': 'proy_cn006_nivel_importancia_id_name', 'nombre_nuevo': 'Importancia'},
        {'nombre_actual': 'proy_cn006_nivel_urgencia_id', 'nombre_nuevo': 'Urgencia ID'},
        {'nombre_actual': 'proy_cn006_nivel_urgencia_id_name', 'nombre_nuevo': 'Urgencia'},
        {'nombre_actual': 'proy_cn006_tamano_id', 'nombre_nuevo': 'Tamaño ID'},
        {'nombre_actual': 'proy_cn006_tamano_id_name', 'nombre_nuevo': 'Tamaño'},
        {'nombre_actual': 'proy_cn006_fecha_creacion_oficial', 'nombre_nuevo': 'Fecha Creación'},
        {'nombre_actual': 'proy_cn006_fecha_inicio_oficial', 'nombre_nuevo': 'Fecha Inicio'},
        {'nombre_actual': 'proy_cn006_fecha_entrega_informatica_oficial', 'nombre_nuevo': 'Fecha Informática'},
        {'nombre_actual': 'proy_cn006_fecha_gerencia_oficial', 'nombre_nuevo': 'Fecha Gerencia'},
        {'nombre_actual': 'proy_cn006_fecha_cierre_oficial', 'nombre_nuevo': 'Fecha Cierre'},
        {'nombre_actual': 'proy_description', 'nombre_nuevo': 'Descripción'},
    ]

    # Extraer solo los campos definidos en la estructura y cambiar sus nombres
    resultado = []
    for proyecto in p_proyectos_info_total:
        if proyecto.get('proy_id_unico', '').strip().upper() == 'X':
            
            nuevo_registro = {}
            for campo in campos_seleccionados:
                nombre_actual = campo['nombre_actual']
                nombre_nuevo = campo['nombre_nuevo']
                valor = proyecto.get(nombre_actual, "")

                #Evitar None en excel
                if valor == None:
                    valor = ""

                # Si es un campo de fecha y es datetime, convertir a date
                if "cn006_fecha" in nombre_actual and isinstance(valor, datetime):
                    valor = valor.strftime('%Y-%m-%d')

                nuevo_registro[nombre_nuevo] = valor

            resultado.append(nuevo_registro)
            msj_debug(f"Proyecto campos seleccionados {proyecto}\n\n", p_tools)

    p_tools.msj_debug(f"Se han seleccionado {len(resultado)} registros con los campos definidos.")
    return resultado

#endregion SELECCIÓN DE DATOS PARA EXCEL
######################################################################################################

######################################################################################################
#region CREAR ARCHIVO EXCEL
def escribir_fecha_excel(sheet, row, col, valor):
    """
    Convierte una fecha a datetime.date si es necesario y la escribe en la celda de Excel.
    Si el valor es None, deja la celda vacía.
    """
    if isinstance(valor, str):  # Si viene como string, convertirlo
        try:
            valor = datetime.strptime(valor, "%Y-%m-%d %H:%M:%S")
        except ValueError:
            valor = None  # Si falla la conversión, dejarlo como None

    sheet.cell(row=row, column=col, value=valor)

from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

def crear_archivo_excel(p_proyectos_info_total, p_tools):
    msj_debug("\n***** Estoy en crear_archivo_excel", p_tools)
    msj_debug(f"Se van a procesar ({len(p_proyectos_info_total)})", p_tools)

    file_path = "cn006_kpi_v030.xlsx"

    # Abrir el archivo Excel existente
    msj_debug("Abriendo archivo existente", p_tools)
    wb = openpyxl.load_workbook(file_path)

    # Verificar si la hoja _odoo_data existe, si no, crearla
    if "_odoo_data" not in wb.sheetnames:
        msj_debug("Creando hoja '_odoo_data'", p_tools)
        ws = wb.create_sheet("_odoo_data")
    else:
        ws = wb["_odoo_data"]

    # Limpiar datos existentes en _odoo_data
    ws.delete_rows(1, ws.max_row)

    #region Encabezados EXCEL
    msj_debug("Escribiendo encabezados en _odoo_data", p_tools)
    headers = list(p_proyectos_info_total[0].keys()) if p_proyectos_info_total else []
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
    #endregion

    #region Datos EXCEL
    msj_debug("Escribiendo datos en _odoo_data", p_tools)
    for row_idx, proyecto in enumerate(p_proyectos_info_total, start=2):
        for col_idx, header in enumerate(headers, start=1):
            valor = proyecto.get(header, None)
            ws.cell(row=row_idx, column=col_idx, value=valor)

            # Formatear celda si es fecha
            if header in COLUMNAS_FECHA:
                ws.cell(row=row_idx, column=col_idx).number_format = "YYYY-MM-DD"
            elif header in COLUMNAS_FECHA_HORA:
                ws.cell(row=row_idx, column=col_idx).number_format = "YYYY-MM-DD HH:MM:SS"
    #endregion

    #region Actualizar tabla _tbl_odoo_data
    msj_debug("Actualizando tabla _tbl_odoo_data", p_tools)
    tabla_existente = None

    # Buscar si existe la tabla
    tbl=""
    msj_debug("\nBuscando tabla _tbl_odoo_data", p_tools)
    

    # tabla_existente = next((tbl for tbl in ws._tables if isinstance(tbl, Table) and tbl.name == "_tbl_odoo_data"), None)

    tbl = ws.tables["_tbl_odoo_data"]
    tabla_existente = tbl
   
    #msj_debug(f"Tabla encontrada: (tabla_existente) ({tabla_existente})", p_tools)

    # Calcular el rango de la tabla
    msj_debug("Calculando rango de la tabla", p_tools)
    max_row = len(p_proyectos_info_total) + 1  # +1 para incluir el encabezado
    max_col = len(headers)
    nuevo_rango = f"A1:{get_column_letter(max_col)}{max_row}"

    if tabla_existente:
        msj_debug(f"Actualizando el rango de la tabla a: {nuevo_rango}", p_tools)
        tabla_existente.ref = nuevo_rango
    else:
        msj_debug("No se encontró una tabla existente, creando una nueva", p_tools)
        tabla = Table(displayName="_tbl_odoo_data", ref=nuevo_rango)
    
        # Establecer estilo para la tabla
        estilo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=True,
        )
        tabla.tableStyleInfo = estilo
        ws.add_table(tabla)
    msj_debug(f"Se actualizó/creó tabla en excel", p_tools)
    #endregion

    # Guardar el archivo sin perder otras hojas
    msj_debug("Guardando archivo Excel actualizado", p_tools)
    wb.save(file_path)
    msj_debug("\n***** FINALIZANDO crear_archivo_excel", p_tools)

#endregion CREAR ARCHIVO EXCEL
######################################################################################################

######################################################################################################
#region JSON PARA INTERFAZ VBA
def convertir_a_json(p_proyectos_info_total, p_tools: cCN006_globales):
    msj_debug(f"\nIniciando convertir_a_json  ({datetime.now().strftime('%Y-%m-%d %H:%M:%S')})", p_tools)

    # Convertir fechas a cadenas
    for item in p_proyectos_info_total:
        for key, value in item.items():
            if isinstance(value, datetime):
                item[key] = value.isoformat()  # Formato: 'YYYY-MM-DDTHH:MM:SS'

    try:
        proyectos_tmp = json.dumps(p_proyectos_info_total, default=lambda x: "" if x is None else x, ensure_ascii=False, indent=4)

        

    except (TypeError, ValueError) as e:
        msj_debug(f"Error al convertir a JSON: {e}", p_tools)
        return None

    msj_debug(f"\nFinalizó convertir_a_json  ({datetime.now().strftime('%Y-%m-%d %H:%M:%S')})", p_tools)
    return proyectos_tmp

def crear_salida (p_proyectos, p_tools: cCN006_globales ):
    msj_debug(f"\nIniciando crear_salida  ({datetime.now().strftime("%Y-%m-%d %H:%M:%S")})", p_tools)

    proyectos_json = ""
    proyectos_json = convertir_a_json(p_proyectos, p_tools)
    data = json.loads(proyectos_json)

    if isinstance(data, list) and len(data) > 0 and isinstance(data[0], dict):
        msj_debug(f"Obtener nombres de campos  ({datetime.now().strftime("%Y-%m-%d %H:%M:%S")})", p_tools)

        # Obtener los nombres de los campos
        fields = data[0].keys()

        # Crear la salida
        output = []

        msj_debug(f"Añadir los nombres de los campos a la salida  ({datetime.now().strftime("%Y-%m-%d %H:%M:%S")})", p_tools)
        # Añadir los nombres de los campos a la salida
        output.append("|NT|".join(fields))

        msj_debug(f"Añadir cada registro a la salida (ciclo)  ({datetime.now().strftime("%Y-%m-%d %H:%M:%S")})", p_tools)

        # Añadir cada registro a la salida
        for record in data:
            for key in record:
                if isinstance(record[key], str):
                    record[key] = record[key].replace('\n', ' ** ')

            row = [str(record.get(field, '')) for field in fields]
            output.append("|NT|".join(row))
    return output
#endregion PRESENTAR JSON PARA INTERFAZ VBA
######################################################################################################

################################################################################################################
# Lógica principal
################################################################################################################
def main(p_ambiente, p_debug) :
    """  ****************************************************************************************
    ESTE SCRIPT GENERA TODA LA INFORMACIÓN DE PROYECTOS, PARTES DE HORAS Y TAREAS ASOCIADAS
    PARA CREAR UN ARCHIVO EXCEL CON TODA LA INFORMACIÓN.

    ES LA BASE PARA ELABORAR LOS KPIs DE PROYECTOS.

    EL DESARROLLO TERMINA ACÁ COMO UN PROTOTIPO Y SE CONTINUA BAJO EL MODELO DE HELPDESK.
    ****************************************************************************************  """

    try:

        #region Instanciar tools
        if p_debug:
            print("010 -  Llamada a inicializar clase")        

        tools = cCN006_globales(p_ambiente, p_debug)
        
        msj_debug("020 -  Retorno de inicializar clase", tools)
        #endregion Instanciar tools

        #region Autenticación con Odoo
        
        msj_debug(f"030 - Llamada autenticar_odoo", tools)

        if (not tools.autenticar_odoo()):
            print (f"(ERROR-010) Error de autenticación")
            return
        
        msj_debug(f"040 - Retorno autenticar_odoo", tools)
        #endregion Autenticación con Odoo
      
        #region Obtener proyectos que se procesarán
        msj_debug(f"050 - Llamando obtener_proyectos", tools)
        proyectos = []
        proyectos = obtener_proyectos(tools)
        msj_debug(f"060 - Regresando obtener_proyectos", tools)

        if len(proyectos) > 0:
            msj_debug(f"Se ecnontraron ({len(proyectos)}) registros de proyectos", tools)
        else:
            msj_debug(f"*** NO SE ENCONTRARON REGISTROS DE PROYECTOS", tools)
        #endregion Obtener proyectos que se procesarán
    
        #region Obtener los partes de horas asociadas a proyectos
        msj_debug(f"070 - Listado de timesheets de todos los proyectos", tools)
        detalles_partes_horas = []
        detalles_partes_horas = obtener_partes_horas(proyectos, tools)
        msj_debug(f"080 - Regresando obtener_partes_horas", tools)
        #endregion Obtener los partes de horas asociadas a proyectos
        
        #region Obtener datos tareas asociadas a partes de horas
        msj_debug(f"090 - Obtener listado de tareas asociadas a partes de horas", tools)
        detalles_partes_horas = obtener_tareas(detalles_partes_horas, tools)
        msj_debug(f"100 - Regresando de listado de tareas asociadas a partes de horas", tools)

        # print(f"Estos son los registros de partes de horas deben tener las tareas asociadas")
        # for detalle in detalles_partes_horas:
        #     print(f"\n{detalle}")
        #endregion Obtener datos tareas asociadas a partes de horas

        #region Unificar proyectos con sus partes de horas para el EXCEL
        msj_debug(f"110 - Unificar datos de proyectos con partes de horas", tools)    
        proyectos_info_total = []
        proyectos_info_total = unificar_datos(proyectos, detalles_partes_horas, tools)
        msj_debug(f"120 - Regresando unificar_datos", tools)

        # Resultado final
        msj_debug(f"En total hay {len(proyectos_info_total)} registros combinados", tools)
        # for fila in proyectos_info_total:
        #     print(fila)
        #endregion Unificar proyectos con sus partes de horas para el EXCEL

        #region Ultimos ajustes antes de crear excel
        msj_debug(f"130 - LLAMANDO ajustes_finales_data", tools)    
        proyectos_info_total = ajustes_finales_data(proyectos_info_total, tools)
        msj_debug(f"140 - REGRESÓ de ajustes_finales_data", tools)    
        #endregion Ultimos ajustes antes de crear excel
        
        #region Seleccionar campos finales
        msj_debug("150 Llamando: Seleccionar campos finales", tools)
        proyectos_info_total = seleccionar_campos_finales(proyectos_info_total, tools)
        msj_debug("160 Regresando: Campos finales seleccionados", tools)
        #endregion Seleccionar campos finales

        #region Generar Datos para interfaz VBA
        msj_debug("Llamando: Generando información para interfaz", tools)
        # crear_archivo_excel(proyectos_info_total, tools)
        salida = []
        salida = crear_salida(proyectos_info_total, tools)
        msj_debug("Regresando: Generando información para interfaz", tools)
        for linea in salida:
            print(linea)
        #endregion Generar Excel

        return detalles_partes_horas

    except Exception as e:
        print(f"\n\n*****   Error general en el proceso:\n>>>>>>>>\n{e}")
        input("Presione cualquier tecla para continuar")
        return []


    # EXCEPCIONES - EXCEPCIONES - EXCEPCIONES - EXCEPCIONES - EXCEPCIONES - EXCEPCIONES - EXCEPCIONES - EXCEPCIONES - EXCEPCIONES - 
    except xmlrpc.client.ProtocolError as error:
        tools.asigna_error(f"Error de protocolo.  {error}|")
        print (f"|Todo Ok: {tools.g_todo_ok}\n*|* conexión: {tools.formatear_datos_conexion()}\n*|* msj: {tools.g_msj}|")            
        input("Presione cualquier tecla para continuar")
        #return g_todo_ok, g_datos_conexion, g_msj
    except xmlrpc.client.Fault as error:
        tools.asigna_error(f"Error de RPC: {error}")
        print (f"|Todo Ok: {tools.g_todo_ok}\n*|* conexión: {tools.formatear_datos_conexion()}\n*|* msj: {tools.g_msj}|")            
        input("Presione cualquier tecla para continuar")
        #return g_todo_ok, g_datos_conexion, g_msj
    except Exception as error:
        tools.asigna_error(f"Error general: {error}")
        print (f"|Todo Ok: {tools.g_todo_ok}\n*|* conexión: {tools.formatear_datos_conexion()}\n*|* msj: {tools.g_msj}|")            
        input("Presione cualquier tecla para continuar")
        #return g_todo_ok, g_datos_conexion, g_msj

################################################################################################################
# FIN - Lógica principal
################################################################################################################


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Script para obtener partes de horas desde Odoo.")

    # Instanciar cCN006_globales antes de usar str2bool
    tmp = cCN006_globales("DESA", False )

    parser.add_argument('--pAmbiente', type=str,          nargs='?', const="DESA", default="DESA", help="Ambiente para ejecutar (DESA/PROD)")
    parser.add_argument('--pDebug',    type=tmp.str2bool, nargs='?', const=True,   default=False,  help="True - Despliega mensajes de seguimiento | False - para correr en producción")

    parser.set_defaults(pAmbiente="DESA", pDebug=False)

    args = parser.parse_args()
    args.pAmbiente = args.pAmbiente.upper()

    del tmp


    # Pasar la instancia globales a main
    main(args.pAmbiente, args.pDebug)
